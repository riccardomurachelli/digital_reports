import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from PIL import Image, ImageTk
import sqlite3
from datetime import datetime, timedelta
import os
import ast
import openpyxl
import locale
import webbrowser

# Configurazione database
def setup_database():
    conn = sqlite3.connect("reports.db")
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER,
            number INTEGER,
            agente1 TEXT,
            agente2 TEXT,
            turno TEXT,
            mezzo TEXT,
            attività TEXT,
            date TEXT,
            infrazioni TEXT,
            identificati TEXT,
            annotazioni TEXT,
            annotazionifinali TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS activities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id INTEGER,
            dalle_ore TEXT,
            alle_ore TEXT,
            codice TEXT,
            via TEXT,
            annotazioni TEXT,
            infrazioni TEXT,
            date TEXT,
            agenti TEXT,
            localita TEXT,
            ore_totali REAL,
            FOREIGN KEY(report_id) REFERENCES reports(id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS identificati (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id INTEGER,
            orario TEXT,
            localita TEXT,
            nominativo TEXT,
            residenza TEXT,
            documento TEXT,
            targa TEXT,
            FOREIGN KEY(report_id) REFERENCES reports(id)
        )
    ''')
    conn.commit()
    conn.close()

setup_database()

# Funzione per ottenere il prossimo numero di report
def get_next_report_number(report_date):
    conn = sqlite3.connect("reports.db")
    cursor = conn.cursor()
    year = datetime.strptime(report_date, "%d/%m/%Y").year
    cursor.execute("SELECT MAX(number) FROM reports WHERE year = ?", (year,))
    result = cursor.fetchone()
    conn.close()
    return (result[0] or 0) + 1

def save_report(agent1, agent2, turno, mezzo, attività, annotazioni, report_date):
    conn = sqlite3.connect("reports.db")
    cursor = conn.cursor()
    year = datetime.now().year
    number = get_next_report_number(report_date)
    # report_date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute(
        "INSERT INTO reports (year, number, agente1, agente2, turno, mezzo, attività, date, annotazioni) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (year, number, agent1, agent2, turno, mezzo, attività, report_date, annotazioni)
    )
    report_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return number, year, report_id

def open_image(root):
    image_window = tk.Toplevel(root)
    image_window.title("Tabella")
    image_window.geometry("600x600")

    try:
        img = Image.open("REDACTED: table")
        img = img.resize((600, 600), Image.LANCZOS)
        img_tk = ImageTk.PhotoImage(img)

        label = tk.Label(image_window, image=img_tk)
        label.image = img_tk
        label.pack()
    except FileNotFoundError:
        messagebox.showerror("Errore", "File 'REDACTED: table' non trovato.")

def start_new_report():
    def get_next_report_number(date):
        conn = sqlite3.connect("reports.db")
        cursor = conn.cursor()
        year = datetime.strptime(date, "%d/%m/%Y").year
        cursor.execute("SELECT MAX(number) FROM reports WHERE year = ?", (year,))
        result = cursor.fetchone()
        conn.close()
        return (result[0] or 0) + 1
    
    def inserisci_attivita(report_id):

        attivita_window = tk.Toplevel(root)
        attivita_window.title("Inserimento Attività")
        attivita_window.geometry("1000x400")
        open_image(attivita_window)

        codici = {
            "REDACTED: codes"
        }
        localita_options = ["REDACTED: locations"]
        columns = ("id", "Agenti", "Dalle ore", "Alle ore", "Codice", "Via", "Località", "Annotazioni", "Infrazioni", "Ore Totali")
        tree = ttk.Treeview(attivita_window, columns=columns, show="headings", selectmode="browse", style="Custom.Treeview")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100, anchor="center")
        tree.pack(fill="both", expand=True)

        style = ttk.Style()
        style.configure("Custom.Treeview.Heading", borderwidth=2, relief="solid")
        style.configure("Custom.Treeview", borderwidth=2, relief="solid")
        style.layout("Custom.Treeview", [("Custom.Treeview.treearea", {"sticky": "nswe"})])

        style.map("Custom.Treeview", background=[("selected", "#cce5ff")], foreground=[("selected", "black")], borderwith=[("selected", 1)])
        style.configure("Custom.Treeview", rowheight=25, borderwidth=1, relief="solid")
        style.configure("Custom.Treeview.Item", borderwidth=1, relief="solid")

        def add_row():
            conn = sqlite3.connect("reports.db")
            cursor = conn.cursor()

            # First, get the available agents from the report
            cursor.execute("SELECT agente1, agente2 FROM reports WHERE id = ?", (report_id,))
            report_agents = cursor.fetchone()
            available_agents = [agent for agent in [report_agents[0], report_agents[1]] if agent != "Nessuno"]

            # Default agent (first available)
            default_agent = available_agents[0] if available_agents else ""

            # Insert a new activity record with some default values
            cursor.execute(
                "INSERT INTO activities (report_id, agenti, dalle_ore, alle_ore, codice, via, localita, annotazioni, infrazioni, date, ore_totali) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                (report_id, default_agent, "08:00", "09:00", "REDACTED: placeholder", "REDACTED: placeholder", "REDACTED: placeholder", "REDACTED: placeholder", "{}", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 1)
            )

            # Get the ID of the newly inserted record
            new_activity_id = cursor.lastrowid

            conn.commit()
            conn.close()

            # Insert the row into the treeview with the database ID as the first value
            tree.insert("", "end", values=(new_activity_id, "", "", "", "", "", "", "", "", ""))

        def calculate_hours(dalle_ore, alle_ore, num_agents=1):
            fmt = '%H:%M'
            tdelta = datetime.strptime(alle_ore, fmt) - datetime.strptime(dalle_ore, fmt)
            return (tdelta.total_seconds() / 3600) * num_agents

        def edit_row(event=None):
            selected_item = tree.selection()
            if not selected_item:
                messagebox.showerror("Errore", "Seleziona una riga da modificare.")
                return
            selected_item = selected_item[0]
            values = tree.item(selected_item, "values")
            edit_window = tk.Toplevel(attivita_window)
            edit_window.title("Modifica Attività")
            edit_window.geometry("400x900")
            conn = sqlite3.connect("reports.db")
            cursor = conn.cursor()
            cursor.execute("SELECT agente1, agente2 FROM reports WHERE id = ?", (report_id,))
            report_agents = cursor.fetchone()
            conn.close()

            available_agents = [agent for agent in [report_agents[0], report_agents[1]] if agent != "Nessuno"]

            tk.Label(edit_window, text="Seleziona agenti:").pack(pady=5)
            agent_vars = {agent: tk.IntVar(value=1 if agent in values[1].split(", ") else 0) for agent in available_agents}
            agent_frame = tk.Frame(edit_window)
            agent_frame.pack(pady=5)

            for agent in available_agents:
                tk.Checkbutton(agent_frame, text=agent, variable=agent_vars[agent]).pack(side=tk.LEFT, padx=5)

            tk.Label(edit_window, text="Dalle ore:").pack(pady=5)
            dalle_ore_entry = ttk.Entry(edit_window, style="Custom.TEntry")
            dalle_ore_entry.insert(0, values[2])
            dalle_ore_entry.pack(pady=5)

            tk.Label(edit_window, text="Alle ore:").pack(pady=5)
            alle_ore_entry = ttk.Entry(edit_window, style="Custom.TEntry")
            alle_ore_entry.insert(0, values[3])
            alle_ore_entry.pack(pady=5)

            tk.Label(edit_window, text="Codice:").pack(pady=5)
            codice_combobox = ttk.Combobox(edit_window, values=list(codici.values()), style="Custom.TCombobox")
            codice_combobox.set(values[4])
            codice_combobox.pack(pady=5)
            codice_combobox.config(state="readonly")

            tk.Label(edit_window, text="Via:").pack(pady=5)
            via_entry = ttk.Entry(edit_window, style="Custom.TEntry")
            via_entry.insert(0, values[5])
            via_entry.pack(pady=5)

            tk.Label(edit_window, text="Località:").pack(pady=5)
            localita_combobox = ttk.Combobox(edit_window, values=localita_options, style="Custom.TCombobox")
            localita_combobox.set(values[6])
            localita_combobox.pack(pady=5)
            localita_combobox.config(state="readonly")

            tk.Label(edit_window, text="Annotazioni:").pack(pady=5)
            annotazioni_entry = tk.Text(edit_window, height=5, width=40, borderwidth=2, relief="solid")
            annotazioni_entry.insert("1.0", values[7])
            annotazioni_entry.pack(pady=5)

            tk.Label(edit_window, text="Seleziona le infrazioni accertate:").pack(pady=5)
            infrazioni = [
                "REDACTED: infractions"
            ]
            infrazioni_vars = {infrazione: tk.IntVar(value=0) for infrazione in infrazioni}

            for infrazione in infrazioni:
                frame = tk.Frame(edit_window)
                frame.pack(anchor="w")
                tk.Label(frame, text=infrazione).pack(side="left")
                tk.Entry(frame, textvariable=infrazioni_vars[infrazione], width=5).pack(side="left")

            def save_changes():
                try:
                    # Raccogli i dati dal form
                    dalle_ore = dalle_ore_entry.get()
                    alle_ore = alle_ore_entry.get()

                    # Validazione degli orari
                    if not dalle_ore or not alle_ore:
                        messagebox.showerror("Errore", "I campi orario non possono essere vuoti.")
                        return

                    try:
                        datetime.strptime(dalle_ore, "%H:%M")
                        datetime.strptime(alle_ore, "%H:%M")
                    except ValueError:
                        messagebox.showerror("Errore", "Formato orario non valido. Usa il formato HH:MM.")
                        return

                    # Calcola le ore totali
                    ore_totali = calculate_hours(dalle_ore, alle_ore)

                    # Continua con le altre operazioni
                    infrazioni_selezionate = {infrazione: var.get() for infrazione, var in infrazioni_vars.items() if var.get() > 0}
                    selected_agents = [agent for agent, var in agent_vars.items() if var.get() == 1]

                    # Prepara i nuovi valori per l'albero
                    new_values = (
                        ", ".join(selected_agents),
                        dalle_ore,
                        alle_ore,
                        codice_combobox.get(),
                        via_entry.get(),
                        localita_combobox.get(),
                        annotazioni_entry.get("1.0", "end").strip(),
                        str(infrazioni_selezionate),
                        ore_totali
                    )

                    # Aggiorna il database
                    record_id = tree.item(selected_item, "values")[0]  # First value is now the database ID
                    conn = sqlite3.connect("reports.db")
                    cursor = conn.cursor()

                    try:
                        cursor.execute(
                            """
                            UPDATE activities
                            SET agenti = ?, dalle_ore = ?, alle_ore = ?, codice = ?, via = ?, localita = ?, annotazioni = ?, infrazioni = ?, ore_totali = ?
                            WHERE id = ?
                            """,
                            (", ".join(selected_agents), dalle_ore, alle_ore, codice_combobox.get(), via_entry.get(), localita_combobox.get(), 
                            annotazioni_entry.get("1.0", "end").strip(), str(infrazioni_selezionate), ore_totali, record_id)
                        )
                        conn.commit()
                        print(f"Aggiornamento attività con ID {record_id} riuscito.")
                    except sqlite3.Error as e:
                        print(f"Errore nell'aggiornamento: {e}")
                        return
                    finally:
                        conn.close()

                    # Aggiorna il treeview con i nuovi valori
                    tree.item(selected_item, values=(record_id, *new_values))

                    # Chiudi la finestra
                    edit_window.destroy()

                except Exception as e:
                    print(f"Errore generico: {e}")

            tk.Button(edit_window, text="Salva", command=save_changes).pack(pady=20)

        add_row_button = tk.Button(attivita_window, text="Aggiungi Riga", command=add_row)
        add_row_button.pack(pady=10)

        save_button = tk.Button(attivita_window, text="Salva Attività", command=lambda: [inserisci_identificati(report_id, attivita_window)])
        save_button.pack(pady=10)

        back_button = tk.Button(attivita_window, text="Indietro", command=lambda: [attivita_window.destroy(), start_new_report()])
        back_button.pack(pady=10)

        tree.bind("<Double-1>", edit_row)

        add_row()

    new_report_window = tk.Toplevel(root)
    new_report_window.title(f"Creazione report {get_next_report_number(datetime.now().strftime('%d/%m/%Y'))}/{datetime.now().year}")
    new_report_window.geometry("400x400")
    tk.Label(new_report_window, text="Il report è per la giornata di oggi?", font=("Arial", 12)).pack(pady=10)
    def select_date():
        date_window = tk.Toplevel(root)
        date_window.title("Seleziona Data")
        date_window.geometry("300x200")

        date_var = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y"))

        tk.Label(date_window, text="Seleziona la data del report:").pack(pady=10)
        date_entry = ttk.Entry(date_window, textvariable=date_var)
        date_entry.pack(pady=10)

        def confirm_date():
            date_window.destroy()
            select_agents(date_var.get())

        tk.Button(date_window, text="Prosegui", command=confirm_date).pack(pady=20)
    def select_agents(report_date):
        new_report_window.destroy()
        agent_window = tk.Toplevel(root)
        agent_window.title(f"Creazione report {get_next_report_number(report_date)}/{datetime.strptime(report_date, '%d/%m/%Y').year}")
        agent_window.geometry("400x400")

        tk.Label(agent_window, text="Seleziona gli agenti:", font=("Arial", 12)).pack(pady=10)

        agent1_var = tk.StringVar(value="Nessuno")
        agent2_var = tk.StringVar(value="Nessuno")

        tk.Label(agent_window, text="Primo agente:").pack(pady=5)
        agent1_menu = ttk.Combobox(agent_window, textvariable=agent1_var, values=["REDACTED: agents"])
        agent1_menu.pack(pady=5)

        tk.Label(agent_window, text="Secondo agente:").pack(pady=5)
        agent2_menu = ttk.Combobox(agent_window, textvariable=agent2_var, values=["REDACTED: agents"])
        agent2_menu.pack(pady=5)

        def select_turno():
            agent_window.destroy()
            turno_window = tk.Toplevel(root)
            turno_window.title("Seleziona Turno")
            turno_window.geometry("300x300")

            turno_var = tk.StringVar()

            tk.Label(turno_window, text="Seleziona il turno di servizio:").pack(pady=10)
            for turno in ["Mattina", "Pomeriggio", "Giornata"]:
                tk.Radiobutton(turno_window, text=turno, variable=turno_var, value=turno).pack(pady=5)

            def select_mezzo():
                turno_window.destroy()
                mezzo_window = tk.Toplevel(root)
                mezzo_window.title("Seleziona Mezzo")
                mezzo_window.geometry("400x300")

                mezzi = []

                def toggle_mezzo(mezzo):
                    if mezzo in mezzi:
                        mezzi.remove(mezzo)
                    else:
                        mezzi.append(mezzo)

                tk.Label(mezzo_window, text="Seleziona i mezzi di trasporto:").pack(pady=10)
                for mezzo in ["REDACTED: vehicles"]:
                    tk.Checkbutton(mezzo_window, text=mezzo, command=lambda m=mezzo: toggle_mezzo(m)).pack(pady=5)

                annotazioni_text = tk.Text(mezzo_window, height=5, width=40)
                tk.Label(mezzo_window, text="Annotazioni:").pack(pady=5)
                annotazioni_text.pack(pady=5)

                def confirm_mezzo():
                    annotazioni = annotazioni_text.get("1.0", "end").strip()
                    mezzo_selezionato = ", ".join(mezzi)

                    mezzo_window.destroy()
                    _, _, report_id = save_report(agent1_var.get(), agent2_var.get(), turno_var.get(), mezzo_selezionato, "", annotazioni, report_date)
                    inserisci_attivita(report_id)
                    tk.Button(mezzo_window, text="Prosegui", command=confirm_mezzo).pack(pady=20)
                    tk.Button(mezzo_window, text="Indietro", command=lambda: [mezzo_window.destroy(), select_turno()]).pack(pady=10)

                tk.Button(turno_window, text="Prosegui", command=select_mezzo).pack(pady=20)
                tk.Button(turno_window, text="Indietro", command=lambda: [turno_window.destroy(), select_agents(report_date)]).pack(pady=10)

            tk.Button(agent_window, text="Prosegui", command=select_turno).pack(pady=20)
            tk.Button(agent_window, text="Indietro", command=lambda: [agent_window.destroy(), select_date()]).pack(pady=10)

    tk.Button(new_report_window, text="Oggi", command=lambda: select_agents(datetime.now().strftime("%d/%m/%Y"))).pack(pady=10)
    tk.Button(new_report_window, text="Altro giorno", command=lambda: select_date()).pack(pady=10)
    def inserisci_identificati(report_id, attivita_window):
        attivita_window.destroy()
        identificati_window = tk.Toplevel(root)
        identificati_window.title("Inserimento Persone Identificate")
        identificati_window.geometry("800x400")

        columns = ("Orario", "Località", "Nominativo", "Residenza", "Documento", "Targa")
        tree = ttk.Treeview(identificati_window, columns=columns, show="headings", selectmode="browse", style="Custom.Treeview")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100, anchor="center")
        tree.pack(fill="both", expand=True)

        style = ttk.Style()
        style.configure("Custom.Treeview.Heading", borderwidth=2, relief="solid")
        style.configure("Custom.Treeview", borderwidth=2, relief="solid")
        style.layout("Custom.Treeview", [("Custom.Treeview.treearea", {"sticky": "nswe"})])

        style.map("Custom.Treeview", background=[("selected", "#cce5ff")], foreground=[("selected", "black")], borderwith=[("selected", 1)])
        style.configure("Custom.Treeview", rowheight=25, borderwidth=1, relief="solid")
        style.configure("Custom.Treeview.Item", borderwidth=1, relief="solid")

        def add_row():
            tree.insert("", "end", values=("", "", "", "", "", ""))

        def save_identificati():
            conn = sqlite3.connect("reports.db")
            cursor = conn.cursor()

            
            for row in tree.get_children():
                values = tree.item(row, "values")
                orario, localita, nominativo, residenza, documento, targa = values
                if orario and localita and nominativo and residenza and documento and targa:
                    cursor.execute(
                        "INSERT INTO identificati (report_id, orario, localita, nominativo, residenza, documento, targa) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (report_id, orario, localita, nominativo, residenza, documento, targa)
                    )

            conn.commit()
            conn.close()
            identificati_window.destroy()

        def edit_row(event=None):
            selected_item = tree.selection()
            if not selected_item:
                messagebox.showerror("Errore", "Seleziona una riga da modificare.")
                return
            selected_item = selected_item[0]
            values = tree.item(selected_item, "values")
            edit_window = tk.Toplevel(identificati_window)
            edit_window.title("Modifica Persona Identificata")
            edit_window.geometry("400x500")

            tk.Label(edit_window, text="Orario:").pack(pady=5)
            orario_entry = ttk.Entry(edit_window, style="Custom.TEntry")
            orario_entry.insert(0, values[0])
            orario_entry.pack(pady=5)

            tk.Label(edit_window, text="Località:").pack(pady=5)
            localita_entry = ttk.Entry(edit_window, style="Custom.TEntry")
            localita_entry.insert(0, values[1])
            localita_entry.pack(pady=5)

            tk.Label(edit_window, text="Nominativo:").pack(pady=5)
            nominativo_entry = ttk.Entry(edit_window, style="Custom.TEntry")
            nominativo_entry.insert(0, values[2])
            nominativo_entry.pack(pady=5)

            tk.Label(edit_window, text="Residenza:").pack(pady=5)
            residenza_entry = ttk.Entry(edit_window, style="Custom.TEntry")
            residenza_entry.insert(0, values[3])
            residenza_entry.pack(pady=5)

            tk.Label(edit_window, text="Documento:").pack(pady=5)
            documento_entry = ttk.Entry(edit_window, style="Custom.TEntry")
            documento_entry.insert(0, values[4])
            documento_entry.pack(pady=5)

            tk.Label(edit_window, text="Targa:").pack(pady=5)
            targa_entry = ttk.Entry(edit_window, style="Custom.TEntry")
            targa_entry.insert(0, values[5])
            targa_entry.pack(pady=5)

            def save_changes():
                new_values = (
                    orario_entry.get(),
                    localita_entry.get(),
                    nominativo_entry.get(),
                    residenza_entry.get(),
                    documento_entry.get(),
                    targa_entry.get()
                )
                
                tree.item(selected_item, values=new_values)
                edit_window.destroy()

            tk.Button(edit_window, text="Salva", command=save_changes).pack(pady=20)

        add_row_button = tk.Button(identificati_window, text="Aggiungi Riga", command=add_row)
        add_row_button.pack(pady=10)

        tree.bind("<Double-1>", edit_row)

        add_row()
        save_button = tk.Button(identificati_window, text="Salva Identificati", command=lambda: [save_identificati(), inserisci_annotazioni_finali(report_id, identificati_window)])
        save_button.pack(pady=10)
        back_button = tk.Button(identificati_window, text="Indietro", command=lambda: [identificati_window.destroy(), inserisci_attivita(report_id)])
        back_button.pack(pady=10)
    def inserisci_annotazioni_finali(report_id, identificati_window):
        identificati_window.destroy()
        annotazioni_finali_window = tk.Toplevel(root)
        annotazioni_finali_window.title("Inserimento Annotazioni Finali")
        annotazioni_finali_window.geometry("400x400")

        annotazioni_finali_text = tk.Text(annotazioni_finali_window, height=10, width=40)
        annotazioni_finali_text.pack(pady=10)

        def save_annotazioni_finali():
            annotazioni_finali = annotazioni_finali_text.get("1.0", "end").strip()
            conn = sqlite3.connect("reports.db")
            cursor = conn.cursor()
            cursor.execute("UPDATE reports SET annotazionifinali = ? WHERE id = ?", (annotazioni_finali, report_id))
            conn.commit()
            conn.close()
            annotazioni_finali_window.destroy()
            messagebox.showinfo("Successo", f"Report {report_id} completato con successo.")

            # Chiedi se stampare il report giornaliero
            if messagebox.askyesno("Stampa Report Giornaliero", "Vuoi stampare il report giornaliero?"):
                stampa_report_giornaliero(report_id, annotazioni_finali)

        tk.Button(annotazioni_finali_window, text="Salva", command=save_annotazioni_finali).pack(pady=10)
        tk.Button(annotazioni_finali_window, text="Indietro", command=lambda: [annotazioni_finali_window.destroy(), inserisci_identificati(report_id, annotazioni_finali_window)]).pack(pady=10)

    def stampa_report_giornaliero(report_id, annotazioni_finali):
        try:
            locale.setlocale(locale.LC_TIME, "it_IT.UTF-8")

            # Carica il file Excel
            workbook = openpyxl.load_workbook("REDACTED: daily report file", keep_vba=True)
            sheet = workbook.active

            # Ottieni la data odierna e il nome del giorno
            today = datetime.now()
            date_str = today.strftime("%d-%m-%Y")
            day_name = today.strftime("%A")

            # Crea un nuovo file basato sul template
            new_filename = f"report_giornaliero_{date_str}.xlsm"
            workbook.save(new_filename)

            # Riapri il nuovo file per modificarlo
            workbook = openpyxl.load_workbook(new_filename, keep_vba=True)
            sheet = workbook.active

            # Scrivi i dati base
            sheet["G5"] = f"{day_name}, {today.strftime('%d/%m/%Y')}"  # Giorno del report con nome
            sheet["J5"] = report_id  # Numero del report

            # Nome degli agenti dal database
            conn = sqlite3.connect("reports.db")
            cursor = conn.cursor()
            cursor.execute("SELECT agente1, agente2, turno FROM reports WHERE id = ?", (report_id,))
            agente_data = cursor.fetchone()
            conn.close()

            if agente_data:
                primo_agente, secondo_agente, turno_servizio = agente_data
                sheet["B6"] = primo_agente if primo_agente else "N/A"
                sheet["E6"] = secondo_agente if secondo_agente else "N/A"
                sheet["J6"] = turno_servizio if turno_servizio else "N/A"

            # Modifica il mezzo utilizzato
            sheet["C8"] = "x"  
            sheet["C9"] = "x"  

            # Annotazioni sul veicolo
            conn = sqlite3.connect("reports.db")
            cursor = conn.cursor()
            cursor.execute("SELECT annotazioni FROM reports WHERE id = ?", (report_id,))
            annotazioni_veicolo = cursor.fetchone()[0]
            conn.close()
            sheet["D8"] = annotazioni_veicolo

            # Ottieni attività dal database
            conn = sqlite3.connect("reports.db")
            cursor = conn.cursor()
            cursor.execute("SELECT dalle_ore, alle_ore, codice, via, localita, annotazioni FROM activities WHERE report_id = ?", (report_id,))
            activities = cursor.fetchall()
            conn.close()

            # Dizionario codici
            codice_map = {
                "REDACTED: codes"
            }

            localita_map = {
                "REDACTED: locations"
            }

            # Scrivi le attività nel report
            start_row = 35
            for i, activity in enumerate(activities):
                dalle_ore, alle_ore, codice, via, localita, annotazioni = activity
                codice_numero = codice_map.get(codice, codice)
                localita_lettera = localita_map.get(localita, localita)
                sheet[f"A{start_row + i}"] = dalle_ore
                sheet[f"B{start_row + i}"] = alle_ore
                sheet[f"C{start_row + i}"] = codice_numero
                sheet[f"D{start_row + i}"] = via
                sheet[f"F{start_row + i}"] = localita_lettera
                sheet[f"G{start_row + i}"] = annotazioni

            # Dati statistici (esempio: Articoli specifici)
            counts = {
                "REDACTED: infractions"
            }

            conn = sqlite3.connect("reports.db")
            cursor = conn.cursor()
            cursor.execute("SELECT infrazioni FROM activities WHERE report_id = ?", (report_id,))
            results = cursor.fetchall()
            conn.close()

            for infrazioni in results:
                try:
                    infrazioni_dict = ast.literal_eval(infrazioni[0]) if infrazioni[0] else {}
                except (ValueError, SyntaxError):
                    infrazioni_dict = {}
                for key in counts.keys():
                    if key in infrazioni_dict:
                        counts[key] += infrazioni_dict[key]

            sheet["C47"] = counts["REDACTED: infractions"]
            sheet["C48"] = counts["REDACTED: infractions"]
            sheet["C49"] = counts["REDACTED: infractions"]
            sheet["F47"] = counts["REDACTED: infractions"]
            sheet["F48"] = counts["REDACTED: infractions"]
            sheet["F49"] = counts["REDACTED: infractions"]
            sheet["I47"] = counts["REDACTED: infractions"]
            sheet["I48"] = counts["REDACTED: infractions"]
            sheet["I49"] = counts["REDACTED: infractions"]

            # Identificazioni
            identificazioni_start_row = 53
            conn = sqlite3.connect("reports.db")
            cursor = conn.cursor()
            cursor.execute("SELECT orario, localita, nominativo, residenza, documento, targa FROM identificati WHERE report_id = ?", (report_id,))
            identificazioni = cursor.fetchall()
            conn.close()

            for i, identificazione in enumerate(identificazioni):
                orario, localita, nominativo, residenza, documento, targa = identificazione
                localita_lettera = localita_map.get(localita, localita)
                sheet[f"A{identificazioni_start_row + i}"] = orario
                sheet[f"B{identificazioni_start_row + i}"] = localita_lettera
                sheet[f"D{identificazioni_start_row + i}"] = nominativo
                sheet[f"F{identificazioni_start_row + i}"] = residenza
                sheet[f"H{identificazioni_start_row + i}"] = documento
                sheet[f"J{identificazioni_start_row + i}"] = targa

            # Annotazioni finali
            sheet["A67"] = annotazioni_finali

            # Salva il file Excel
            workbook.save(new_filename)
            messagebox.showinfo("Successo", f"Report giornaliero salvato come {new_filename}!")

        except Exception as e:
            messagebox.showerror("Errore", f"Errore durante la stampa del report: {e}")
    
        
def view_progress():
    progress_window = tk.Toplevel(root)
    progress_window.title("Visualizza Andamento")
    progress_window.geometry("800x600")

    tk.Label(progress_window, text="Andamento Report", font=("Arial", 16)).pack(pady=10)

    tk.Label(progress_window, text="Per quanto tempo vuoi vedere l'andamento?").pack(pady=10)

    time_frame_var = tk.StringVar(value="Settimanale")

    time_frames = ["Settimanale", "Mensile", "Trimestrale", "Annuale"]
    time_frame_menu = ttk.Combobox(progress_window, textvariable=time_frame_var, values=time_frames)
    time_frame_menu.pack(pady=10)

    columns = ("Attività", "Numero di Occorrenze", "Infrazioni", "Ore Totali")
    report_list = ttk.Treeview(progress_window, columns=columns, show="headings")
    for col in columns:
        report_list.heading(col, text=col)
        report_list.column(col, width=150, anchor="center")
    report_list.pack(fill="both", expand=True)

    def load_reports():
        for row in report_list.get_children():
            report_list.delete(row)

        time_frame = time_frame_var.get()
        conn = sqlite3.connect("reports.db")
        cursor = conn.cursor()

        if time_frame == "Settimanale":
            start_date = (datetime.now() - timedelta(days=7)).date()
        elif time_frame == "Mensile":
            start_date = (datetime.now() - timedelta(days=30)).date()
        elif time_frame == "Trimestrale":
            start_date = (datetime.now() - timedelta(days=90)).date()
        elif time_frame == "Annuale":
            start_date = (datetime.now() - timedelta(days=365)).date()

        query = """
        SELECT a.codice, COUNT(*), GROUP_CONCAT(a.infrazioni), SUM(a.ore_totali)
        FROM activities a
        JOIN reports r ON a.report_id = r.id
        WHERE date(r.date) >= ? 
        GROUP BY a.codice
        """
        cursor.execute(query, (start_date.isoformat(),))
        rows = cursor.fetchall()
        conn.close()

        for row in rows:
            report_list.insert("", "end", values=row)

        def on_time_frame_change(*args):
            load_reports()

        time_frame_var.trace_add("write", on_time_frame_change)

    load_reports()

    def export_to_excel():
        time_frame = time_frame_var.get()
        file_name = f"andamento_{time_frame.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        template_path = "REDACTED: template for reports"  # File modello
        try:
            workbook = openpyxl.load_workbook(template_path)
        except FileNotFoundError:
            print("Errore: File modello non trovato.")
            return

        sheet = workbook.active

        sheet["H6"] = datetime.now().strftime("%Y-%m-%d")  # Data di esportazione
        sheet["J6"] = datetime.now().strftime("%H:%M:%S")  # Ora di esportazione
        sheet["H7"] = time_frame  # Intervallo di esportazione

        try:
            conn = sqlite3.connect("reports.db")
            cursor = conn.cursor()
        except sqlite3.Error as e:
            print(f"Errore di connessione al database: {e}")
            return

        counts = {
            "REDACTED: infractions" : 0
        }

        if time_frame == "Settimanale":
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        elif time_frame == "Mensile":
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        elif time_frame == "Trimestrale":
            start_date = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
        elif time_frame == "Annuale":
            start_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')

        cursor.execute("SELECT infrazioni FROM activities WHERE date >= ?", (start_date,))
        results = cursor.fetchall()
        for infrazioni in results:
            try:
                infrazioni_dict = ast.literal_eval(infrazioni[0]) if infrazioni[0] else {}
            except (ValueError, SyntaxError):
                infrazioni_dict = {}
            for key in counts.keys():
                if key in infrazioni_dict:
                    counts[key] += infrazioni_dict[key]

        counts_mapping = {
            "REDACETD: infractions": "C11",
        }

        for key, cell in counts_mapping.items():
            sheet[cell] = counts[key]
            activities = [
                "REDACTED: activities"
            ]

            activity_cells = {
                "REDACTED: activities": "C15",
            }


        # Conteggio delle attività e delle ore totali
        activity_counts = {activity: 0 for activity in activities}
        activity_hours = {activity: 0 for activity in activities}
        for activity in activities:
            cursor.execute(f"SELECT COUNT(*), IFNULL(SUM(ore_totali), 0) FROM activities WHERE codice LIKE ? AND date >= '{start_date}'", (f"%{activity}%",))
            result = cursor.fetchone()
            activity_counts[activity] = result[0]
            activity_hours[activity] = result[1]

        # Scrittura dei conteggi delle attività e delle ore nelle celle
        for activity, cell in activity_cells.items():
            sheet[cell] = activity_hours[activity]

        conn.close()

        # Salva il file aggiornato
        try:
            workbook.save(file_name)
            messagebox.showinfo("Esportazione completata", f"Andamento esportato in {file_name}")
        except Exception as e:
            print(f"Errore durante il salvataggio del file: {e}")

    tk.Button(progress_window, text="Esporta in Excel", command=export_to_excel).pack(pady=10)

    load_reports()

def print_daily_sheet():
    def select_sheet_type():
        sheet_window = tk.Toplevel(root)
        sheet_window.title("Seleziona Tipo di Foglio")
        sheet_window.geometry("300x200")

        tk.Label(sheet_window, text="Seleziona il tipo di foglio da stampare:").pack(pady=10)

        def print_sheet(sheet_type):
            sheet_window.destroy()
            if sheet_type == "multilocalita":
                file_name = "REDACTED: multi-location sheet"
            else:
                file_name = "REDACTED: single-location sheet"
            result_label.config(text=f"{file_name} pronto per la stampa.")
            os.startfile(file_name)

        tk.Button(sheet_window, text="Foglio Multilocalità", command=lambda: print_sheet("multilocalita")).pack(pady=10)
        tk.Button(sheet_window, text="Foglio Semplice", command=lambda: print_sheet("semplice")).pack(pady=10)

    select_sheet_type()

def stampa_report_giornaliero(report_id, annotazioni_finali):
    try:
        locale.setlocale(locale.LC_TIME, "it_IT.utf8")

        # Carica il file Excel
        workbook = openpyxl.load_workbook("REDACTED: File name daily report", keep_vba=True)
        sheet = workbook.active

        # Ottieni la data odierna e il nome del giorno
        conn = sqlite3.connect("reports.db")
        cursor = conn.cursor()
        cursor.execute("SELECT date FROM reports WHERE id = ?", (report_id,))
        report_date = cursor.fetchone()[0]
        conn.close()

        report_date_obj = datetime.strptime(report_date, "%d/%m/%Y")
        date_str = report_date_obj.strftime("%d-%m-%Y")
        day_name = report_date_obj.strftime("%A")

        # Crea un nuovo file basato sul template
        new_filename = f"report_giornaliero_{date_str}.xlsm"
        workbook.save(new_filename)

        # Riapri il nuovo file per modificarlo
        workbook = openpyxl.load_workbook(new_filename, keep_vba=True)
        sheet = workbook.active

        # Scrivi i dati base
        sheet["G5"] = f"{day_name}, {date_str}"  # Giorno del report con nome
        sheet["J5"] = report_id  # Numero del report

        # Nome degli agenti dal database
        conn = sqlite3.connect("reports.db")
        cursor = conn.cursor()
        cursor.execute("SELECT agente1, agente2, turno FROM reports WHERE id = ?", (report_id,))
        agente_data = cursor.fetchone()
        conn.close()

        if agente_data:
            primo_agente, secondo_agente, turno_servizio = agente_data
            sheet["B6"] = primo_agente if primo_agente else "N/A"
            sheet["E6"] = secondo_agente if secondo_agente else "N/A"
            sheet["J6"] = turno_servizio if turno_servizio else "N/A"

        # Modifica il mezzo utilizzato
        sheet["C8"] = "x"  
        sheet["C9"] = "x"  

        # Annotazioni sul veicolo
        conn = sqlite3.connect("reports.db")
        cursor = conn.cursor()
        cursor.execute("SELECT annotazioni FROM reports WHERE id = ?", (report_id,))
        annotazioni_veicolo = cursor.fetchone()[0]
        conn.close()
        sheet["D8"] = annotazioni_veicolo

        # Ottieni attività dal database
        conn = sqlite3.connect("reports.db")
        cursor = conn.cursor()
        cursor.execute("SELECT dalle_ore, alle_ore, codice, via, localita, annotazioni FROM activities WHERE report_id = ?", (report_id,))
        activities = cursor.fetchall()
        conn.close()

        # Dizionario codici
        codice_map = {
            "REDACTED: codes"
        }

        localita_map = {
            "REDACTED: locations"
        }

        # Scrivi le attività nel report
        start_row = 35
        for i, activity in enumerate(activities):
            dalle_ore, alle_ore, codice, via, localita, annotazioni = activity
            codice_numero = codice_map.get(codice, codice)
            localita_lettera = localita_map.get(localita, localita)
            sheet[f"A{start_row + i}"] = dalle_ore
            sheet[f"B{start_row + i}"] = alle_ore
            sheet[f"C{start_row + i}"] = codice_numero
            sheet[f"D{start_row + i}"] = via
            sheet[f"F{start_row + i}"] = localita_lettera
            sheet[f"G{start_row + i}"] = annotazioni

        # Dati statistici (esempio: Articoli specifici)
        counts = {
            "REDACTED: infractions" : 0
        }

        conn = sqlite3.connect("reports.db")
        cursor = conn.cursor()
        cursor.execute("SELECT infrazioni FROM activities WHERE report_id = ?", (report_id,))
        results = cursor.fetchall()
        conn.close()

        for infrazioni in results:
            try:
                infrazioni_dict = ast.literal_eval(infrazioni[0]) if infrazioni[0] else {}
            except (ValueError, SyntaxError):
                infrazioni_dict = {}
            for key in counts.keys():
                if key in infrazioni_dict:
                    counts[key] += infrazioni_dict[key]

        sheet["C47"] = counts["REDACTED: infractions"]
        sheet["C48"] = counts["REDACTED: infractions"]
        sheet["C49"] = counts["REDACTED: infractions"]
        sheet["F47"] = counts["REDACTED: infractions"]
        sheet["F48"] = counts["REDACTED: infractions"]
        sheet["F49"] = counts["REDACTED: infractions"]
        sheet["I47"] = counts["REDACTED: infractions"]
        sheet["I48"] = counts["REDACTED: infractions"]
        sheet["I49"] = counts["REDACTED: infractions"]

        # Identificazioni
        identificazioni_start_row = 53
        conn = sqlite3.connect("reports.db")
        cursor = conn.cursor()
        cursor.execute("SELECT orario, localita, nominativo, residenza, documento, targa FROM identificati WHERE report_id = ?", (report_id,))
        identificazioni = cursor.fetchall()
        conn.close()

        for i, identificazione in enumerate(identificazioni):
            orario, localita, nominativo, residenza, documento, targa = identificazione
            localita_lettera = localita_map.get(localita, localita)
            sheet[f"A{identificazioni_start_row + i}"] = orario
            sheet[f"B{identificazioni_start_row + i}"] = localita_lettera
            sheet[f"D{identificazioni_start_row + i}"] = nominativo
            sheet[f"F{identificazioni_start_row + i}"] = residenza
            sheet[f"H{identificazioni_start_row + i}"] = documento
            sheet[f"J{identificazioni_start_row + i}"] = targa

        # Annotazioni finali
        sheet["A67"] = annotazioni_finali

        # Salva il file Excel
        workbook.save(new_filename)
        messagebox.showinfo("Successo", f"Report giornaliero salvato come {new_filename}!")

    except Exception as e:
        messagebox.showerror("Errore", f"Errore durante la stampa del report: {e}")

def edit_report():
    def select_report():
        report_window = tk.Toplevel(root)
        report_window.title("Seleziona Report")
        report_window.geometry("400x400")

        tk.Label(report_window, text="Seleziona il report da modificare:", font=("Arial", 12)).pack(pady=10)

        columns = ("ID", "Data", "Agenti", "Turno", "Mezzo", "Annotazioni")
        report_list = ttk.Treeview(report_window, columns=columns, show="headings")
        for col in columns:
            report_list.heading(col, text=col)
            report_list.column(col, width=100, anchor="center")
        report_list.pack(fill="both", expand=True)

        conn = sqlite3.connect("reports.db")
        cursor = conn.cursor()
        cursor.execute("SELECT id, date, agente1, agente2, turno, mezzo, annotazioni FROM reports")
        rows = cursor.fetchall()
        conn.close()

        for row in rows:
            report_list.insert("", "end", values=row)

        def on_report_select(event):
            selected_item = report_list.selection()
            if not selected_item:
                messagebox.showerror("Errore", "Seleziona un report.")
                return
            selected_item = selected_item[0]
            report_id = report_list.item(selected_item, "values")[0]

            action_window = tk.Toplevel(report_window)
            action_window.title("Azione Report")
            action_window.geometry("300x200")

            def modify_report():
                action_window.destroy()
                report_window.destroy()
                modify_report_window(report_id)

            def print_report():
                action_window.destroy()
                report_window.destroy()
                stampa_report_giornaliero(report_id, "")

            def delete_report():
                action_window.destroy()
                if messagebox.askyesno("Conferma Eliminazione", "Sei sicuro di voler eliminare questo report?"):
                    conn = sqlite3.connect("reports.db")
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM reports WHERE id = ?", (report_id,))
                    cursor.execute("DELETE FROM activities WHERE report_id = ?", (report_id,))
                    cursor.execute("DELETE FROM identificati WHERE report_id = ?", (report_id,))
                    conn.commit()
                    conn.close()
                    messagebox.showinfo("Successo", "Report eliminato con successo.")
                    report_list.delete(selected_item)

            tk.Button(action_window, text="Modifica", command=modify_report).pack(pady=10)
            tk.Button(action_window, text="Stampa", command=print_report).pack(pady=10)
            tk.Button(action_window, text="Elimina", command=delete_report).pack(pady=10)

        report_list.bind("<Double-1>", on_report_select)

    def modify_report_window(report_id):
        edit_report_window = tk.Toplevel(root)
        edit_report_window.title(f"Modifica Report {report_id}")
        edit_report_window.geometry("800x1000")

        tk.Label(edit_report_window, text=f"Modifica Report {report_id}", font=("Arial", 16)).pack(pady=10)

        conn = sqlite3.connect("reports.db")
        cursor = conn.cursor()
        cursor.execute("SELECT date, agente1, agente2, turno, mezzo, annotazioni FROM reports WHERE id = ?", (report_id,))
        report_data = cursor.fetchone()
        conn.close()

        date_var = tk.StringVar(value=report_data[0])
        agent1_var = tk.StringVar(value=report_data[1])
        agent2_var = tk.StringVar(value=report_data[2])
        turno_var = tk.StringVar(value=report_data[3])
        mezzo_var = tk.StringVar(value=report_data[4])
        annotazioni_var = tk.StringVar(value=report_data[5])

        tk.Label(edit_report_window, text="Data:").pack(pady=5)
        date_entry = ttk.Entry(edit_report_window, textvariable=date_var)
        date_entry.pack(pady=5)

        tk.Label(edit_report_window, text="Primo Agente:").pack(pady=5)
        agent1_entry = ttk.Entry(edit_report_window, textvariable=agent1_var)
        agent1_entry.pack(pady=5)

        tk.Label(edit_report_window, text="Secondo Agente:").pack(pady=5)
        agent2_entry = ttk.Entry(edit_report_window, textvariable=agent2_var)
        agent2_entry.pack(pady=5)

        tk.Label(edit_report_window, text="Turno:").pack(pady=5)
        turno_entry = ttk.Entry(edit_report_window, textvariable=turno_var)
        turno_entry.pack(pady=5)

        tk.Label(edit_report_window, text="Mezzo:").pack(pady=5)
        mezzo_entry = ttk.Entry(edit_report_window, textvariable=mezzo_var)
        mezzo_entry.pack(pady=5)

        tk.Label(edit_report_window, text="Annotazioni:").pack(pady=5)
        annotazioni_entry = tk.Text(edit_report_window, height=5, width=40)
        annotazioni_entry.insert("1.0", annotazioni_var.get())
        annotazioni_entry.pack(pady=5)

        tk.Label(edit_report_window, text="Attività", font=("Arial", 12)).pack(pady=10)

        columns = ("ID", "Agenti", "Dalle Ore", "Alle Ore", "Codice", "Via", "Località", "Annotazioni", "Infrazioni", "Ore Totali")
        tree = ttk.Treeview(edit_report_window, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100, anchor="center")
        tree.pack(fill="both", expand=True)

        conn = sqlite3.connect("reports.db")
        cursor = conn.cursor()
        cursor.execute("SELECT id, agenti, dalle_ore, alle_ore, codice, via, localita, annotazioni, infrazioni, ore_totali FROM activities WHERE report_id = ?", (report_id,))
        rows = cursor.fetchall()
        conn.close()

        for row in rows:
            tree.insert("", "end", values=row)

        def add_row():
            tree.insert("", "end", values=("", "", "", "", "", "", "", "", "", ""))

        def save_changes():
            conn = sqlite3.connect("reports.db")
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE reports SET date = ?, agente1 = ?, agente2 = ?, turno = ?, mezzo = ?, annotazioni = ? WHERE id = ?",
                (date_var.get(), agent1_var.get(), agent2_var.get(), turno_var.get(), mezzo_var.get(), annotazioni_entry.get("1.0", "end").strip(), report_id)
            )
            for row in tree.get_children():
                values = tree.item(row, "values")
                if values[0]:
                    cursor.execute(
                        "UPDATE activities SET agenti = ?, dalle_ore = ?, alle_ore = ?, codice = ?, via = ?, localita = ?, annotazioni = ?, infrazioni = ?, ore_totali = ? WHERE id = ?",
                        (values[1], values[2], values[3], values[4], values[5], values[6], values[7], values[8], values[9], values[0])
                    )
            conn.commit()
            conn.close()
            messagebox.showinfo("Successo", "Modifiche salvate con successo.")
            edit_report_window.destroy()

        add_row_button = tk.Button(edit_report_window, text="Aggiungi Riga", command=add_row)
        add_row_button.pack(pady=10)

        save_button = tk.Button(edit_report_window, text="Salva Modifiche", command=save_changes)
        save_button.pack(pady=10)

    select_report()
    
# Creazione della finestra principale
root = tk.Tk()
root.title("REDACTED: Title")
root.geometry("500x400")
# root.iconphoto(False, tk.PhotoImage(file='icon.png'))

# Etichetta del titolo principale
tk.Label(root, text="REDACTED: Title", font=("Arial", 16)).pack(pady=20)

# Menu principale con pulsanti
tk.Button(root, text="Inizia Nuovo Report", command=start_new_report, width=30).pack(pady=10)
tk.Button(root, text="Visualizza Andamento", command=view_progress, width=30).pack(pady=10)
tk.Button(root, text="Stampa Foglio Giornaliero", command=print_daily_sheet, width=30).pack(pady=10)
tk.Button(root, text="Modifica Report", command=edit_report, width=30).pack(pady=10)

# Etichetta per mostrare il risultato delle azioni
result_label = tk.Label(root, text="", font=("Arial", 12))
result_label.pack(pady=20)

# Funzioni per aprire i link
def open_website(url):
    webbrowser.open_new(url)
    


# Etichetta con testo su una sola riga
label = tk.Label(
    root,
    text="Codice scritto con il 🩷 da Riccardo Murachelli | Open-Source su GitHub",
    font=("Arial", 10, "italic"),
    fg="blue",
    cursor="hand2"
)
label.pack(pady=10)

# Collegamenti ai link
def on_label_click(event):
    if event.x < 275:  # Posizione approssimativa del primo link
        open_website("http://www.riccardomurachelli.it")
    else:
        open_website("https://github.com/riccardomurachelli/digital_reports")

label.bind("<Button-1>", on_label_click)    

# Avvio della finestra
root.mainloop()
